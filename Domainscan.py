import concurrent.futures
import nmap
import openpyxl
from tqdm import tqdm
from ipaddress import ip_network, ip_address
import socket

# Open the text file containing the IP addresses and subnets
with open('ip_ranges.txt') as f:
    ip_ranges = f.read().splitlines()

# Create a list of IP addresses to scan
ip_list = []
for ip_range in ip_ranges:
    try:
        network = ip_network(ip_range)
        ip_list.extend([str(ip) for ip in network])
    except ValueError:
        ip_list.append(ip_range)

# Create an Nmap PortScanner object
nm = nmap.PortScanner()

# Create an Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Set the column headings for the worksheet
worksheet['A1'] = 'IP Address'
worksheet['B1'] = 'Domain Name'
worksheet['C1'] = 'Port 80'
worksheet['D1'] = 'Port 443'

# Define a function to scan a single IP address
def scan_ip(ip):
    # Check if the input IP address is a valid domain name
    try:
        ip_address(socket.gethostbyname(ip))
        is_ip = True
    except socket.gaierror:
        is_ip = False

    # If the input is a domain name, get the IP address associated with it
    if not is_ip:
        try:
            ip = socket.gethostbyname(ip)
        except socket.gaierror:
            pass

    # Scan ports 80 and 443 for the IP address
    nm.scan(ip, '80,443')

    # Get the status of port 80 and 443 for the IP address
    try:
        port_80_status = nm[ip]['tcp'][80]['state']
    except KeyError:
        port_80_status = 'unknown'
    try:
        port_443_status = nm[ip]['tcp'][443]['state']
    except KeyError:
        port_443_status = 'unknown'

    # Get the domain name for the IP address
    try:
        domain_name = socket.gethostbyaddr(ip)[0]
    except socket.herror:
        domain_name = 'unknown'

    # Return the IP address or domain name, and port status as a tuple
    return (ip if is_ip else domain_name, domain_name, port_80_status, port_443_status)

# Use multi-threading to scan IP addresses in parallel
with concurrent.futures.ThreadPoolExecutor(max_workers=50) as executor:
    results = list(tqdm(executor.map(scan_ip, ip_list), total=len(ip_list)))

# Write the results to the Excel worksheet
for i, result in enumerate(results):
    worksheet.cell(row=i+2, column=1).value = result[0]
    worksheet.cell(row=i+2, column=2).value = result[1]
    worksheet.cell(row=i+2, column=3).value = result[2]
    worksheet.cell(row=i+2, column=4).value = result[3]

# Save the Excel workbook
workbook.save('port_scan_results202.xlsx')
